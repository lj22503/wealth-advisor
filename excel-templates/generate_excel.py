#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
生成 Excel 模板文件
使用：python generate_excel_templates.py
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

# 颜色定义
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
DATA_FONT = Font(size=10)
CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)

# 边框
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)


def create_excel_template(filename, headers, sample_data, column_widths, required_cols=None):
    """创建 Excel 模板文件"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "模板"
    
    # 设置表头
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER
    
    # 设置示例数据
    for row_idx, row_data in enumerate(sample_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = DATA_FONT
            cell.alignment = LEFT_ALIGN if col_idx in required_cols else CENTER_ALIGN
            cell.border = THIN_BORDER
            
            # 标记必填列为红色字体
            if required_cols and col_idx in required_cols:
                cell.font = Font(color="FF0000", size=10)
    
    # 设置列宽
    for col_idx, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    
    # 设置行高
    ws.row_dimensions[1].height = 25
    for row_idx in range(2, len(sample_data) + 2):
        ws.row_dimensions[row_idx].height = 20
    
    # 冻结首行
    ws.freeze_panes = "A2"
    
    # 添加说明工作表
    ws_help = wb.create_sheet(title="填写说明")
    ws_help.cell(row=1, column=1, value="字段说明")
    ws_help.cell(row=1, column=2, value="是否必填")
    ws_help.cell(row=1, column=3, value="格式要求")
    ws_help.cell(row=1, column=4, value="示例")
    
    help_headers = ["字段说明", "是否必填", "格式要求", "示例"]
    for col, header in enumerate(help_headers, 1):
        cell = ws_help.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER
    
    # 保存文件
    wb.save(filename)
    print(f"✅ 已生成：{filename}")


def generate_all_templates():
    """生成所有模板"""
    output_dir = os.path.dirname(os.path.abspath(__file__))
    
    # 1. 客户数据模板
    customer_headers = ["客户姓名", "手机号", "邮箱", "风险等级", "总资产"]
    customer_sample = [
        ["张三", "13800138000", "zhangsan@example.com", "C3", "500000"],
        ["李四", "13900139000", "lisi@example.com", "C2", "300000"],
        ["王五", "13700137000", "", "C4", "800000"]
    ]
    create_excel_template(
        os.path.join(output_dir, "客户数据模板.xlsx"),
        customer_headers,
        customer_sample,
        [15, 15, 25, 12, 12],
        required_cols=[1, 4]  # 客户姓名、风险等级必填
    )
    
    # 2. 持仓数据模板
    holding_headers = ["客户 ID", "基金代码", "基金名称", "持有份额", "买入价格", "买入日期"]
    holding_sample = [
        ["client_001", "000001", "华夏成长混合", "10000", "1.5000", "2024-01-15"],
        ["client_001", "000002", "易方达蓝筹精选", "5000", "2.3000", "2024-02-20"],
        ["client_002", "000003", "招商中证白酒", "8000", "1.8000", "2024-03-10"]
    ]
    create_excel_template(
        os.path.join(output_dir, "持仓数据模板.xlsx"),
        holding_headers,
        holding_sample,
        [15, 12, 20, 12, 12, 12],
        required_cols=[1, 2, 4, 5]  # 客户 ID、基金代码、持有份额、买入价格必填
    )
    
    # 3. 基金数据模板
    fund_headers = ["基金代码", "基金名称", "基金类型", "风险等级", "单位净值", "净值日期", "基金经理", "成立日期"]
    fund_sample = [
        ["000001", "华夏成长混合", "混合型", "C3", "2.5678", "2024-03-17", "王亚伟", "2015-06-18"],
        ["000002", "易方达蓝筹精选", "混合型", "C4", "3.1234", "2024-03-17", "张坤", "2018-09-20"],
        ["000003", "招商中证白酒", "指数型", "C5", "1.9876", "2024-03-17", "侯昊", "2016-05-12"]
    ]
    create_excel_template(
        os.path.join(output_dir, "基金数据模板.xlsx"),
        fund_headers,
        fund_sample,
        [12, 20, 12, 12, 12, 12, 12, 12],
        required_cols=[1, 2, 3, 4, 5, 6]  # 除基金经理和成立日期外都必填
    )
    
    print("\n🎉 所有模板已生成完成！")
    print("\n📂 文件位置:")
    print(f"   - {output_dir}/客户数据模板.xlsx")
    print(f"   - {output_dir}/持仓数据模板.xlsx")
    print(f"   - {output_dir}/基金数据模板.xlsx")
    print("\n💡 使用说明:")
    print("   1. 下载对应的 Excel 模板")
    print("   2. 按照示例填写数据（红色字体为必填项）")
    print("   3. 在飞书应用中导入 Excel 文件")


if __name__ == "__main__":
    try:
        import openpyxl
        generate_all_templates()
    except ImportError:
        print("❌ 缺少依赖：openpyxl")
        print("请运行：pip install openpyxl")
        exit(1)
